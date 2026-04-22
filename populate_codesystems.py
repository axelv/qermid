# /// script
# requires-python = ">=3.11"
# dependencies = ["openpyxl"]
# ///
"""Extract real codes from DCD Excel files and populate CodeSystem stubs in codesystems.fsh."""

import re
from collections import defaultdict
from pathlib import Path

import openpyxl

FSH_DIR = Path("input/fsh")
EXCEL_DIR = Path("dcd-excel")


# ---------------------------------------------------------------------------
# 1. Parse existing codesystems.fsh
# ---------------------------------------------------------------------------

def load_codesystems() -> dict[str, dict]:
    """Load all CodeSystem definitions from the original (git HEAD) codesystems.fsh.

    Always reads from git HEAD to ensure we start from the full set of 299 definitions,
    even if a previous run already removed some.

    Returns {fsh_name: {"id": ..., "title": ..., "description": ...}}.
    """
    import subprocess

    try:
        result = subprocess.run(
            ["git", "show", "HEAD:input/fsh/codesystems.fsh"],
            capture_output=True, text=True,
        )
        text = result.stdout if result.returncode == 0 and result.stdout else None
    except FileNotFoundError:
        text = None
    if text is None:
        text = (FSH_DIR / "codesystems.fsh").read_text()

    cs_map: dict[str, dict] = {}
    for m in re.finditer(
        r"CodeSystem:\s+(\S+)\nId:\s+(.*)\nTitle:\s+\"(.*)\"\nDescription:\s+\"(.*)\"",
        text,
    ):
        cs_map[m.group(1)] = {
            "id": m.group(2).strip(),
            "title": m.group(3),
            "description": m.group(4),
        }
    return cs_map


# ---------------------------------------------------------------------------
# 2. Name matching: Excel code list type -> CodeSystem FSH name
# ---------------------------------------------------------------------------

def build_title_to_cs_map(cs_map: dict[str, dict]) -> dict[str, str]:
    """Build a lookup from normalized title to CS FSH name.

    The CS title is the original code list name (e.g. "BILLING_CODE_172 code + label").
    We normalize it to match what the Excel TYPE column contains.
    """
    title_to_cs: dict[str, str] = {}
    for fsh_name, info in cs_map.items():
        title = info["title"]
        title_to_cs[title.upper().strip()] = fsh_name
        # Also index without "code + label" suffix
        clean = re.sub(r"\s*code\s*\+\s*label\s*", "", title, flags=re.IGNORECASE).strip()
        if clean != title:
            title_to_cs[clean.upper().strip()] = fsh_name
    return title_to_cs


def match_codelist_to_cs(
    codelist_name: str,
    title_to_cs: dict[str, str],
) -> str | None:
    """Match an Excel code list name to a CodeSystem FSH name.

    Tries multiple normalization strategies.
    """
    name = codelist_name.strip()
    # Normalize: collapse multiple spaces, strip
    norm = " ".join(name.split()).upper()

    # Direct match on title
    if norm in title_to_cs:
        return title_to_cs[norm]

    # Try with parenthetical number removed
    no_parens = re.sub(r"\s*\(\d+\)", "", norm).strip()
    if no_parens in title_to_cs:
        return title_to_cs[no_parens]

    # Try with "(code + label)" added
    with_cl = f"{norm} CODE + LABEL"
    if with_cl in title_to_cs:
        return title_to_cs[with_cl]

    no_parens_cl = f"{no_parens} CODE + LABEL"
    if no_parens_cl in title_to_cs:
        return title_to_cs[no_parens_cl]

    # Try matching titles that have extra spaces (e.g., "CEMENT_NAME _81" vs "CEMENT_NAME_81")
    # Normalize both sides by removing all spaces
    norm_nospace = norm.replace(" ", "")
    for key, cs_name in title_to_cs.items():
        if key.replace(" ", "") == norm_nospace:
            return cs_name

    # Try prefix matching
    # e.g., Excel has "BILLING_CODE_172" and CS title is "BILLING_CODE_172 code + label"
    for key, cs_name in title_to_cs.items():
        if key.startswith(norm):
            return cs_name

    return None


# ---------------------------------------------------------------------------
# 3. FSH code escaping
# ---------------------------------------------------------------------------

def fsh_code(code: str) -> str:
    """Format a code value for FSH. Quote if it contains special characters."""
    code = str(code).strip()
    # Remove embedded double quotes — FSH quoted codes don't support escaped quotes
    code = code.replace('"', "")
    # Collapse multiple spaces to single space
    code = " ".join(code.split())
    if re.match(r"^[A-Za-z0-9_.]+$", code):
        return f"#{code}"
    return f'#"{code}"'


def fsh_string(s: str) -> str:
    """Escape a string for FSH double-quoted context."""
    return str(s).replace("\\", "\\\\").replace('"', '\\"')


def _verhoeff_check(num: str) -> bool:
    """Validate a number string using the Verhoeff checksum algorithm.

    SNOMED CT concept IDs use a Verhoeff check digit as the last digit.
    """
    d = [
        [0, 1, 2, 3, 4, 5, 6, 7, 8, 9],
        [1, 2, 3, 4, 0, 6, 7, 8, 9, 5],
        [2, 3, 4, 0, 1, 7, 8, 9, 5, 6],
        [3, 4, 0, 1, 2, 8, 9, 5, 6, 7],
        [4, 0, 1, 2, 3, 9, 5, 6, 7, 8],
        [5, 9, 8, 7, 6, 0, 4, 3, 2, 1],
        [6, 5, 9, 8, 7, 1, 0, 4, 3, 2],
        [7, 6, 5, 9, 8, 2, 1, 0, 4, 3],
        [8, 7, 6, 5, 9, 3, 2, 1, 0, 4],
        [9, 8, 7, 6, 5, 4, 3, 2, 1, 0],
    ]
    p = [
        [0, 1, 2, 3, 4, 5, 6, 7, 8, 9],
        [1, 5, 7, 6, 2, 8, 3, 0, 9, 4],
        [5, 8, 0, 3, 7, 9, 6, 1, 4, 2],
        [8, 9, 1, 6, 0, 4, 3, 5, 2, 7],
        [9, 4, 5, 3, 1, 2, 6, 8, 7, 0],
        [4, 2, 8, 6, 5, 7, 3, 9, 0, 1],
        [2, 7, 9, 3, 8, 0, 6, 4, 1, 5],
        [7, 0, 4, 6, 9, 1, 3, 2, 5, 8],
    ]
    c = 0
    for i, digit in enumerate(reversed(num)):
        c = d[c][p[i % 8][int(digit)]]
    return c == 0


def is_snomed_code(code: str) -> bool:
    """Check if a code value is a valid SNOMED CT concept ID.

    SNOMED CT IDs are 6-18 digit numbers with a valid Verhoeff check digit.
    """
    s = str(code).strip()
    if not re.match(r"^\d{6,18}$", s):
        return False
    return _verhoeff_check(s)


# ---------------------------------------------------------------------------
# 4. Extract from dedicated code list sheets
# ---------------------------------------------------------------------------

CodeMap = dict[str, list[tuple[str, str]]]  # {codelist_type: [(code, display), ...]}


def extract_heart_valves_codelists() -> CodeMap:
    """Extract from heart-valves.xlsx CodeLists sheet."""
    path = EXCEL_DIR / "heart-valves.xlsx"
    if not path.exists():
        return {}

    wb = openpyxl.load_workbook(path)
    if "CodeLists" not in wb.sheetnames:
        return {}

    ws = wb["CodeLists"]
    result: CodeMap = defaultdict(list)

    for r in range(2, ws.max_row + 1):
        type_name = ws.cell(r, 2).value  # TYPE column
        code_value = ws.cell(r, 5).value  # CODE_VALUE column
        label_nl = ws.cell(r, 8).value  # LABEL_NL column
        label_en = ws.cell(r, 7).value  # LABEL_EN column (fallback)

        if not type_name or code_value is None:
            continue

        code = str(code_value).strip()
        display = str(label_nl or label_en or code).strip()
        result[str(type_name).strip()].append((code, display))

    return dict(result)


def extract_pacemaker_nominative_lists() -> CodeMap:
    """Extract from pacemakers.xlsx Nominative-Limitative lists sheet."""
    path = EXCEL_DIR / "pacemakers.xlsx"
    if not path.exists():
        return {}

    wb = openpyxl.load_workbook(path)
    sheet_name = "Nominative-Limitative lists"
    if sheet_name not in wb.sheetnames:
        return {}

    ws = wb[sheet_name]
    result: CodeMap = defaultdict(list)

    for r in range(2, ws.max_row + 1):
        type_name = ws.cell(r, 2).value  # TYPE column
        code_value = ws.cell(r, 3).value  # CODE_VALUE column
        label_en = ws.cell(r, 4).value  # LABEL_EN column
        label_nl = ws.cell(r, 5).value  # LABEL_NL
        label_fr = ws.cell(r, 6).value  # LABEL_FR

        if not type_name or code_value is None:
            continue

        code = str(code_value).strip()
        display = str(label_en or label_nl or label_fr or code).strip()
        result[str(type_name).strip()].append((code, display))

    return dict(result)


def extract_cement_name() -> CodeMap:
    """Extract from hip.xlsx Codelist CEMENT_NAME sheet."""
    path = EXCEL_DIR / "hip.xlsx"
    if not path.exists():
        return {}

    wb = openpyxl.load_workbook(path)
    sheet_name = "Codelist CEMENT_NAME"
    if sheet_name not in wb.sheetnames:
        return {}

    ws = wb[sheet_name]
    result: CodeMap = defaultdict(list)

    for r in range(2, ws.max_row + 1):
        codelist = ws.cell(r, 1).value  # CODELIST column
        code_value = ws.cell(r, 2).value  # CODE_VALUE column
        label_en = ws.cell(r, 3).value  # LABEL_EN column
        label_nl = ws.cell(r, 4).value  # LABEL_NL
        label_fr = ws.cell(r, 5).value  # LABEL_FR

        if not codelist or code_value is None:
            continue

        code = str(code_value).strip()
        display = str(label_en or label_nl or label_fr or code).strip()
        result[str(codelist).strip()].append((code, display))

    return dict(result)


# ---------------------------------------------------------------------------
# 5. Extract inline values from Fields sheets
# ---------------------------------------------------------------------------

# (excel_file, sheet_name, fsh_file)
SHEET_MAP: list[tuple[str, str, str]] = [
    # Orthopride Hip
    ("hip.xlsx", "Fields_Primo-implantatie", "qermid-orthopride-hip-primo-implantatie.fsh"),
    ("hip.xlsx", "Fields_Revisie", "qermid-orthopride-hip-revisie.fsh"),
    ("hip.xlsx", "Fields_Resectie", "qermid-orthopride-hip-resectie.fsh"),
    # Orthopride Knee
    ("knee.xlsx", "Fields_Primo-implantatie", "qermid-orthopride-knee-primo-implantatie.fsh"),
    ("knee.xlsx", "Fields_Revisie", "qermid-orthopride-knee-revisie.fsh"),
    ("knee.xlsx", "Fields_Resectie", "qermid-orthopride-knee-resectie.fsh"),
    # Orthopride Total Femur
    ("total-femur.xlsx", "Flds_Totale femur - primo", "qermid-orthopride-total-femur-totale-femur---primo.fsh"),
    ("total-femur.xlsx", "Flds_Totale femur - revisie", "qermid-orthopride-total-femur-totale-femur---revisie.fsh"),
    ("total-femur.xlsx", "Fields_Totale femur - resectie", "qermid-orthopride-total-femur-totale-femur---resectie.fsh"),
    # Pacemakers
    ("pacemakers.xlsx", "Primo-implantation", "qermid-pacemakers-primo-implantation.fsh"),
    ("pacemakers.xlsx", "Primo Follow-up T1", "qermid-pacemakers-primo-follow-up-t1.fsh"),
    ("pacemakers.xlsx", "Remplacement", "qermid-pacemakers-remplacement.fsh"),
    ("pacemakers.xlsx", "Ajout remplacement electrode", "qermid-pacemakers-ajout-remplacement-electrode.fsh"),
    ("pacemakers.xlsx", "Explantation", "qermid-pacemakers-explantation.fsh"),
    ("pacemakers.xlsx", "Follow-up SITI", "qermid-pacemakers-follow-up-siti.fsh"),
    ("pacemakers.xlsx", "Remplacement follow-up T2-T6", "qermid-pacemakers-remplacement-follow-up-t2-t6.fsh"),
    # Angio
    ("angio.xlsx", "Fields-Hospitalisatie", "qermid-angio-fields-hospitalisatie.fsh"),
    ("angio.xlsx", "Fields-Hospitalisatie met PCI", "qermid-angio-fields-hospitalisatie-met-pci.fsh"),
    ("angio.xlsx", "Fields-Hospitalisatie met FFR", "qermid-angio-fields-hospitalisatie-met-ffr.fsh"),
    ("angio.xlsx", "Fields-Hospitalisatie FFR_PCI", "qermid-angio-fields-hospitalisatie-ffr-pci.fsh"),
    ("angio.xlsx", "Fields-Follow-up na PCI", "qermid-angio-fields-follow-up-na-pci.fsh"),
    # Heart Valves
    ("heart-valves.xlsx", "Implantation ", "qermid-heart-valves-implantation-.fsh"),
    ("heart-valves.xlsx", "Follow-up", "qermid-heart-valves-follow-up.fsh"),
]


def _classify_header(header: str) -> str | None:
    """Classify a column header into a semantic role.

    Returns one of: "tech", "nl_label", "fr_label", "expected_result",
    "codelist_values", "codelist_name", or None.
    """
    h = header.strip().lower()

    # Most specific patterns first to avoid false matches

    # Technical name column
    if "technical name" in h:
        return "tech"

    # Codelist values (must check before codelist_name since "codelist" is a substring)
    if "codelist values" in h:
        return "codelist_values"

    # Codelist name / code list name and code values
    if "code list name" in h or h in ("codelist", "codelists", "codelist name"):
        return "codelist_name"

    # FR labels — check before NL to avoid "fields (fr)" matching as NL
    if any(p in h for p in ("fields_fr", "fields (fr)", "label_fr", "vertaling fr")):
        return "fr_label"
    if h == "expected result (fr)":
        return "fr_label"

    # NL labels — explicit patterns only
    if any(p in h for p in ("translation fields", "labels nl")):
        return "nl_label"

    # Expected result (used as NL label for inline values on ortho sheets)
    if "expected result" in h:
        return "expected_result"

    return None


def detect_columns(ws: openpyxl.worksheet.worksheet.Worksheet) -> tuple[int, dict[str, int]]:
    """Detect column positions from the header row.

    Returns (header_row, {role: col_number}).
    Scans rows 5-25 looking for a row that contains a "technical name" header.
    """
    for r in range(5, 25):
        for c in range(1, 20):
            val = ws.cell(r, c).value
            if val and "technical name" in str(val).lower():
                # Found the header row — now map all columns
                cols: dict[str, int] = {}
                for col in range(1, min(ws.max_column + 1, 20)):
                    header = ws.cell(r, col).value
                    if not header:
                        continue
                    role = _classify_header(str(header))
                    if role and role not in cols:
                        cols[role] = col
                return r, cols

    return 22, {}


def extract_inline_values(ws: openpyxl.worksheet.worksheet.Worksheet) -> CodeMap:
    """Extract inline code values from any Fields sheet using header-driven column detection."""
    header_row, cols = detect_columns(ws)
    data_start = header_row + 1

    tech_col = cols.get("tech")
    codelist_name_col = cols.get("codelist_name")
    codelist_values_col = cols.get("codelist_values")
    # For NL display labels: prefer explicit NL label column, fall back to expected_result
    nl_label_col = cols.get("nl_label") or cols.get("expected_result")

    if not tech_col or not codelist_name_col:
        return {}

    # Determine where code values appear in inline rows:
    # - Sheets with a separate "Codelist values" column: code values are there
    # - Sheets without it: code values appear in the "codelist_name" column on inline rows
    inline_code_col = codelist_values_col or codelist_name_col

    result: CodeMap = defaultdict(list)
    current_codelist: str | None = None

    for r in range(data_start, ws.max_row + 1):
        tech = ws.cell(r, tech_col).value

        if tech:
            # Field definition row — check for a codelist reference
            cl = ws.cell(r, codelist_name_col).value
            if cl and not str(cl).upper().startswith("LINK"):
                current_codelist = str(cl).strip().split("\n")[0].strip()

                # On ortho sheets, the field definition row itself may contain
                # the first code value (in the codelist_values column)
                if codelist_values_col:
                    code_val = ws.cell(r, codelist_values_col).value
                    label = ws.cell(r, nl_label_col).value if nl_label_col else None
                    if (
                        code_val is not None
                        and label
                        and "list" not in str(label).lower()
                    ):
                        result[current_codelist].append(
                            (str(code_val).strip(), str(label).strip())
                        )
            else:
                current_codelist = None

        elif current_codelist:
            # Inline value row — try NL label col first, fall back to others
            code_val = ws.cell(r, inline_code_col).value
            label = ws.cell(r, nl_label_col).value if nl_label_col else None

            # If no NL label found, try expected_result col (different from nl_label)
            if not label and cols.get("expected_result") and cols.get("expected_result") != nl_label_col:
                label = ws.cell(r, cols["expected_result"]).value

            if code_val is not None and label:
                code = str(code_val).strip()
                display = str(label).strip()
                if code and display:
                    result[current_codelist].append((code, display))
            elif not any(ws.cell(r, c).value for c in range(1, min(ws.max_column + 1, 15))):
                current_codelist = None

    return dict(result)


def extract_all_inline_values() -> CodeMap:
    """Extract inline code values from all Fields sheets across all Excel files."""
    merged: CodeMap = defaultdict(list)
    wb_cache: dict[str, openpyxl.Workbook] = {}

    for excel_file, sheet_name, _ in SHEET_MAP:
        if excel_file not in wb_cache:
            path = EXCEL_DIR / excel_file
            if not path.exists():
                continue
            wb_cache[excel_file] = openpyxl.load_workbook(path)

        wb = wb_cache[excel_file]
        if sheet_name not in wb.sheetnames:
            print(f"  WARNING: Sheet '{sheet_name}' not found in {excel_file}")
            continue

        ws = wb[sheet_name]
        values = extract_inline_values(ws)

        for codelist_name, codes in values.items():
            merged[codelist_name].extend(codes)

    # Deduplicate codes within each code list (by code value)
    deduped: CodeMap = {}
    for name, codes in merged.items():
        seen: set[str] = set()
        unique: list[tuple[str, str]] = []
        for code, display in codes:
            if code not in seen:
                seen.add(code)
                unique.append((code, display))
        deduped[name] = unique

    return deduped


# ---------------------------------------------------------------------------
# 6. Match extracted codes to CodeSystems and write output
# ---------------------------------------------------------------------------

def classify_codes(
    codes: list[tuple[str, str]],
) -> tuple[list[tuple[str, str]], list[tuple[str, str]]]:
    """Split a code list into SNOMED codes and custom codes.

    Returns (snomed_codes, custom_codes).
    """
    snomed: list[tuple[str, str]] = []
    custom: list[tuple[str, str]] = []
    for code, display in codes:
        if is_snomed_code(code):
            snomed.append((code, display))
        else:
            custom.append((code, display))
    return snomed, custom


def match_all_codes(
    all_codes: CodeMap,
    title_to_cs: dict[str, str],
) -> tuple[dict[str, list[tuple[str, str]]], list[str]]:
    """Match Excel codelist names to CS FSH names and return merged codes."""
    cs_codes: dict[str, list[tuple[str, str]]] = {}
    unmatched: list[str] = []

    for codelist_name, codes in all_codes.items():
        cs_name = match_codelist_to_cs(codelist_name, title_to_cs)
        if cs_name:
            if cs_name not in cs_codes:
                cs_codes[cs_name] = []
            existing = {c for c, _ in cs_codes[cs_name]}
            for code, display in codes:
                if code not in existing:
                    cs_codes[cs_name].append((code, display))
                    existing.add(code)
        else:
            unmatched.append(codelist_name)

    return cs_codes, sorted(set(unmatched))


def write_codesystems(
    cs_map: dict[str, dict],
    cs_codes: dict[str, list[tuple[str, str]]],
) -> dict[str, str]:
    """Write updated codesystems.fsh.

    CodeSystems that only contain SNOMED codes are removed entirely.
    Mixed CodeSystems keep only the custom (non-SNOMED) codes.

    Returns a dict {cs_fsh_name: category} where category is
    "all_snomed", "mixed", "custom_only", or "stub".
    """
    lines = ["// CodeSystems for QERMID registers", ""]
    categories: dict[str, str] = {}

    for fsh_name, info in cs_map.items():
        codes = cs_codes.get(fsh_name, [])
        if not codes:
            # Stub — keep as-is
            lines.append(f"CodeSystem: {fsh_name}")
            lines.append(f"Id: {info['id']}")
            lines.append(f"Title: \"{fsh_string(info['title'])}\"")
            lines.append(f"Description: \"{fsh_string(info['description'])}\"")
            lines.append('* #unknown "Unknown" "Unknown value"')
            lines.append("")
            categories[fsh_name] = "stub"
            continue

        snomed, custom = classify_codes(codes)

        if snomed and not custom:
            # All SNOMED — don't emit a CodeSystem at all
            categories[fsh_name] = "all_snomed"
            continue

        if snomed and custom:
            # Mixed — only keep custom codes in the CodeSystem
            categories[fsh_name] = "mixed"
            emit_codes = custom
        else:
            # All custom
            categories[fsh_name] = "custom_only"
            emit_codes = custom

        lines.append(f"CodeSystem: {fsh_name}")
        lines.append(f"Id: {info['id']}")
        lines.append(f"Title: \"{fsh_string(info['title'])}\"")
        lines.append(f"Description: \"{fsh_string(info['description'])}\"")
        for code, display in emit_codes:
            lines.append(f"* {fsh_code(code)} \"{fsh_string(display)}\"")
        lines.append("")

    (FSH_DIR / "codesystems.fsh").write_text("\n".join(lines))
    return categories


def load_valuesets() -> dict[str, dict]:
    """Load all ValueSet definitions from the original (git HEAD) valuesets.fsh.

    Returns {fsh_name: {"id": ..., "title": ..., "description": ..., "cs_ref": ...}}.
    """
    import subprocess

    try:
        result = subprocess.run(
            ["git", "show", "HEAD:input/fsh/valuesets.fsh"],
            capture_output=True, text=True,
        )
        text = result.stdout if result.returncode == 0 and result.stdout else None
    except FileNotFoundError:
        text = None
    if text is None:
        text = (FSH_DIR / "valuesets.fsh").read_text()

    vs_map: dict[str, dict] = {}
    for m in re.finditer(
        r"ValueSet:\s+(\S+)\nId:\s+(.*)\nTitle:\s+\"(.*)\"\nDescription:\s+\"(.*)\"\n\* include codes from system (\S+)",
        text,
    ):
        vs_map[m.group(1)] = {
            "id": m.group(2).strip(),
            "title": m.group(3),
            "description": m.group(4),
            "cs_ref": m.group(5),
        }
    return vs_map


def write_valuesets(
    vs_map: dict[str, dict],
    cs_codes: dict[str, list[tuple[str, str]]],
    categories: dict[str, str],
) -> None:
    """Write updated valuesets.fsh.

    - All-SNOMED: ValueSet includes specific SCT codes (no CodeSystem reference)
    - Mixed: ValueSet includes from both $SCT (specific codes) and the custom CodeSystem
    - Custom-only / stub: ValueSet includes from the CodeSystem as before
    """
    lines = ["// ValueSets for QERMID registers", ""]

    for vs_name, info in vs_map.items():
        cs_ref = info["cs_ref"]
        cat = categories.get(cs_ref, "stub")

        lines.append(f"ValueSet: {vs_name}")
        lines.append(f"Id: {info['id']}")
        lines.append(f"Title: \"{fsh_string(info['title'])}\"")
        lines.append(f"Description: \"{fsh_string(info['description'])}\"")

        codes = cs_codes.get(cs_ref, [])
        snomed, custom = classify_codes(codes) if codes else ([], [])

        if cat == "all_snomed":
            # Include specific SNOMED codes
            for code, display in snomed:
                lines.append(f"* $SCT{fsh_code(code)} \"{fsh_string(display)}\"")
        elif cat == "mixed":
            # Include SNOMED codes + all codes from custom CodeSystem
            for code, display in snomed:
                lines.append(f"* $SCT{fsh_code(code)} \"{fsh_string(display)}\"")
            lines.append(f"* include codes from system {cs_ref}")
        else:
            # Custom-only or stub — include from CodeSystem as before
            lines.append(f"* include codes from system {cs_ref}")

        lines.append("")

    (FSH_DIR / "valuesets.fsh").write_text("\n".join(lines))


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    # Load existing CodeSystem definitions
    cs_map = load_codesystems()
    print(f"Loaded {len(cs_map)} CodeSystem definitions")

    title_to_cs = build_title_to_cs_map(cs_map)

    # Collect codes from all sources
    all_codes: CodeMap = defaultdict(list)

    # Source 1: Dedicated code list sheets
    print("\nExtracting from dedicated code list sheets...")

    hv_codes = extract_heart_valves_codelists()
    print(f"  Heart Valves CodeLists: {len(hv_codes)} types, {sum(len(v) for v in hv_codes.values())} codes")
    for name, codes in hv_codes.items():
        all_codes[name].extend(codes)

    pm_codes = extract_pacemaker_nominative_lists()
    print(f"  Pacemakers Nominative-Limitative: {len(pm_codes)} types, {sum(len(v) for v in pm_codes.values())} codes")
    for name, codes in pm_codes.items():
        all_codes[name].extend(codes)

    cement_codes = extract_cement_name()
    print(f"  Cement Name: {len(cement_codes)} types, {sum(len(v) for v in cement_codes.values())} codes")
    for name, codes in cement_codes.items():
        all_codes[name].extend(codes)

    # Source 2: Inline values from Fields sheets
    print("\nExtracting inline values from Fields sheets...")
    inline_codes = extract_all_inline_values()
    print(f"  Inline: {len(inline_codes)} code lists, {sum(len(v) for v in inline_codes.values())} codes")

    # Merge inline codes — inline NL labels take priority over dedicated sheet EN labels
    for name, codes in inline_codes.items():
        if name not in all_codes:
            all_codes[name] = codes
        else:
            # Override display labels from dedicated sheets with inline NL labels,
            # and add any codes not already present
            existing = {c: i for i, (c, _) in enumerate(all_codes[name])}
            for code, display in codes:
                if code in existing:
                    # Replace EN label with NL label
                    all_codes[name][existing[code]] = (code, display)
                else:
                    existing[code] = len(all_codes[name])
                    all_codes[name].append((code, display))

    print(f"\nTotal: {len(all_codes)} unique code lists, {sum(len(v) for v in all_codes.values())} codes")

    # Match codes to CodeSystems
    cs_codes, unmatched = match_all_codes(dict(all_codes), title_to_cs)

    # Write codesystems.fsh (SNOMED-only CodeSystems are removed)
    print("\nWriting codesystems.fsh...")
    categories = write_codesystems(cs_map, cs_codes)

    # Write valuesets.fsh (SNOMED codes referenced via $SCT)
    print("Writing valuesets.fsh...")
    vs_map = load_valuesets()
    write_valuesets(vs_map, cs_codes, categories)

    # Summary
    from collections import Counter
    counts = Counter(categories.values())
    print(f"\n{'='*60}")
    print(f"CodeSystem categories:")
    print(f"  All SNOMED (CS removed, VS references $SCT):  {counts.get('all_snomed', 0)}")
    print(f"  Mixed (CS has custom only, VS has $SCT + CS):  {counts.get('mixed', 0)}")
    print(f"  Custom only (CS + VS as before):               {counts.get('custom_only', 0)}")
    print(f"  Stubs (unchanged):                             {counts.get('stub', 0)}")

    if unmatched:
        print(f"\nUnmatched Excel code lists ({len(unmatched)}):")
        for name in unmatched:
            print(f"  {name}")

    stub_names = [n for n, c in categories.items() if c == "stub"]
    if stub_names:
        print(f"\nRemaining stub CodeSystems ({len(stub_names)}):")
        for name in stub_names:
            print(f"  {name} ({cs_map[name]['title']})")


if __name__ == "__main__":
    main()
