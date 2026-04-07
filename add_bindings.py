"""Extract field-to-codelist mappings from DCD Excel files and add ValueSet bindings to FSH logical models."""

import re
from pathlib import Path

import openpyxl


FSH_DIR = Path("input/fsh")
EXCEL_DIR = Path("dcd-excel")


def load_valueset_ids() -> set[str]:
    """Load all ValueSet FSH names from valuesets.fsh."""
    text = (FSH_DIR / "valuesets.fsh").read_text()
    return set(re.findall(r"^ValueSet:\s+(\S+)", text, re.MULTILINE))


def normalize_codelist(name: str, vs_ids: set[str]) -> str | None:
    """Normalize an Excel codelist name to a matching ValueSet FSH name.

    Handles multi-line cell values like "NIC_TPE_44 (44)\\nprefilled" and
    "(code + label)" suffixes. Tries multiple normalization strategies.
    """
    if not name or str(name).upper().startswith("LINK:"):
        return None

    # Parse multi-line values: extract suffixes (prefilled, code + label)
    lines = str(name).strip().split("\n")
    base_line = lines[0].strip()
    suffixes = {s.strip().lower() for s in lines[1:] if s.strip()}

    codelabel = "code + label" in suffixes
    prefilled = "prefilled" in suffixes

    # Skip LINK: references (case-insensitive, with optional space after colon)
    if re.match(r"link\s*:", base_line, flags=re.IGNORECASE):
        return None

    # Remove inline "(code + label)" too
    if re.search(r"\(code\s*\+\s*label\)", base_line, flags=re.IGNORECASE):
        codelabel = True
        base_line = re.sub(r"\s*\(code\s*\+\s*label\)", "", base_line, flags=re.IGNORECASE).strip()

    # Extract parenthetical number
    m = re.search(r"\((\d+)\)", base_line)
    num = m.group(1) if m else ""
    base = re.sub(r"\s*\(\d+\)", "", base_line).strip()

    # Generate candidate base forms
    b_underscore = base.lower().replace(" ", "_")  # spaces → underscores
    b_nospace = base.lower().replace(" ", "")  # spaces removed

    # Build suffix combinations to try
    suffix_combos: list[str] = []
    if prefilled and codelabel:
        suffix_combos.extend(["prefilled", "codelabel", ""])
    elif prefilled:
        suffix_combos.extend(["prefilled", ""])
    elif codelabel:
        suffix_combos.extend(["codelabel", ""])
    else:
        suffix_combos.append("")

    def try_candidate(c: str) -> str | None:
        return c if c in vs_ids else None

    for sfx in suffix_combos:
        for b in [b_underscore, b_nospace]:
            # With number: various patterns
            if num:
                # e.g. nic_tpe_4444prefilled_VS
                for c in [
                    f"{b}{num}{sfx}_VS",           # base+num+suffix
                    f"{b}_{num}{sfx}_VS",           # base_num+suffix
                    f"{b}{num}{num}{sfx}_VS",       # base+num+num+suffix (doubled)
                    f"{b}_{num}{num}{sfx}_VS",      # base_num+num+suffix
                ]:
                    result = try_candidate(c)
                    if result:
                        return result

            # Without number
            for c in [
                f"{b}{sfx}_VS",
                f"{b}_{sfx}_VS" if sfx else None,
            ]:
                if c:
                    result = try_candidate(c)
                    if result:
                        return result

    return None


def find_header_row(ws: openpyxl.worksheet.worksheet.Worksheet, col: int, keyword: str = "technical") -> int:
    """Find the header row by searching for a keyword in the given column."""
    for r in range(5, 25):
        val = ws.cell(r, col).value
        if val and keyword in str(val).lower():
            return r + 1
    return 22  # fallback


def build_orthopride_config(
    excel_file: str, sheet_name: str, fsh_file: str,
) -> tuple[str, str, int, int, int, int, str]:
    """Build config for an Orthopride-style sheet (col4=tech, col5=type, col8=codelist)."""
    wb = openpyxl.load_workbook(EXCEL_DIR / excel_file)
    ws = wb[sheet_name]
    hr = find_header_row(ws, 4, "technical")
    return (excel_file, sheet_name, 4, 5, 8, hr, fsh_file)


def build_pacemaker_config(
    excel_file: str, sheet_name: str, fsh_file: str,
) -> tuple[str, str, int, int, int, int, str]:
    """Build config for a Pacemaker/TAVI-style sheet (col1=tech, col3=type, col5=codelist)."""
    wb = openpyxl.load_workbook(EXCEL_DIR / excel_file)
    ws = wb[sheet_name]
    hr = find_header_row(ws, 1, "technical")
    return (excel_file, sheet_name, 1, 3, 5, hr, fsh_file)


def build_angio_config(
    excel_file: str, sheet_name: str, fsh_file: str,
) -> tuple[str, str, int, int, int, int, str]:
    """Build config for Angio-style sheet (col3=tech, col4=type, col6=codelist)."""
    wb = openpyxl.load_workbook(EXCEL_DIR / excel_file)
    ws = wb[sheet_name]
    hr = find_header_row(ws, 3, "technical")
    return (excel_file, sheet_name, 3, 4, 6, hr, fsh_file)


# Explicit mapping: (excel_file, sheet_name, fsh_file, layout)
# layout: "ortho" = Orthopride-style, "pm" = Pacemaker/ANGIO/TAVI-style
SHEET_MAP: list[tuple[str, str, str, str]] = [
    # Orthopride Hip
    ("hip.xlsx", "Fields_Primo-implantatie", "qermid-orthopride-hip-primo-implantatie.fsh", "ortho"),
    ("hip.xlsx", "Fields_Revisie", "qermid-orthopride-hip-revisie.fsh", "ortho"),
    ("hip.xlsx", "Fields_Resectie", "qermid-orthopride-hip-resectie.fsh", "ortho"),
    # Orthopride Knee
    ("knee.xlsx", "Fields_Primo-implantatie", "qermid-orthopride-knee-primo-implantatie.fsh", "ortho"),
    ("knee.xlsx", "Fields_Revisie", "qermid-orthopride-knee-revisie.fsh", "ortho"),
    ("knee.xlsx", "Fields_Resectie", "qermid-orthopride-knee-resectie.fsh", "ortho"),
    # Orthopride Total Femur
    ("total-femur.xlsx", "Flds_Totale femur - primo", "qermid-orthopride-total-femur-totale-femur---primo.fsh", "ortho"),
    ("total-femur.xlsx", "Flds_Totale femur - revisie", "qermid-orthopride-total-femur-totale-femur---revisie.fsh", "ortho"),
    ("total-femur.xlsx", "Fields_Totale femur - resectie", "qermid-orthopride-total-femur-totale-femur---resectie.fsh", "ortho"),
    # Pacemakers
    ("pacemakers.xlsx", "Primo-implantation", "qermid-pacemakers-primo-implantation.fsh", "pm"),
    ("pacemakers.xlsx", "Primo Follow-up T1", "qermid-pacemakers-primo-follow-up-t1.fsh", "pm"),
    ("pacemakers.xlsx", "Remplacement", "qermid-pacemakers-remplacement.fsh", "pm"),
    ("pacemakers.xlsx", "Ajout remplacement electrode", "qermid-pacemakers-ajout-remplacement-electrode.fsh", "pm"),
    ("pacemakers.xlsx", "Explantation", "qermid-pacemakers-explantation.fsh", "pm"),
    ("pacemakers.xlsx", "Follow-up SITI", "qermid-pacemakers-follow-up-siti.fsh", "pm"),
    ("pacemakers.xlsx", "Remplacement follow-up T2-T6", "qermid-pacemakers-remplacement-follow-up-t2-t6.fsh", "pm"),
    # Angio (different column layout: col3=tech, col4=type, col6=codelist)
    ("angio.xlsx", "Fields-Hospitalisatie", "qermid-angio-fields-hospitalisatie.fsh", "angio"),
    ("angio.xlsx", "Fields-Hospitalisatie met PCI", "qermid-angio-fields-hospitalisatie-met-pci.fsh", "angio"),
    ("angio.xlsx", "Fields-Hospitalisatie met FFR", "qermid-angio-fields-hospitalisatie-met-ffr.fsh", "angio"),
    ("angio.xlsx", "Fields-Hospitalisatie FFR_PCI", "qermid-angio-fields-hospitalisatie-ffr-pci.fsh", "angio"),
    ("angio.xlsx", "Fields-Follow-up na PCI", "qermid-angio-fields-follow-up-na-pci.fsh", "angio"),
    # Heart Valves
    ("heart-valves.xlsx", "Implantation ", "qermid-heart-valves-implantation-.fsh", "pm"),
    ("heart-valves.xlsx", "Follow-up", "qermid-heart-valves-follow-up.fsh", "pm"),
]


def extract_mappings(
    config: tuple[str, str, int, int, int, int, str],
    vs_ids: set[str],
) -> dict[str, str]:
    """Extract {tech_name: valueset_name} from one Excel sheet."""
    excel_file, sheet_name, tc, typc, clc, hr, _ = config
    wb = openpyxl.load_workbook(EXCEL_DIR / excel_file)
    ws = wb[sheet_name]

    mappings: dict[str, str] = {}
    for r in range(hr, ws.max_row + 1):
        tech = ws.cell(r, tc).value
        typ = ws.cell(r, typc).value
        cl = ws.cell(r, clc).value
        if not tech or not typ:
            continue
        if "list" not in str(typ).lower():
            continue
        if not cl:
            continue
        vs = normalize_codelist(str(cl), vs_ids)
        if vs:
            mappings[str(tech).strip()] = vs

    return mappings


def add_bindings_to_fsh(fsh_file: str, mappings: dict[str, str]) -> int:
    """Add ValueSet bindings to code fields in an FSH file.

    Returns number of bindings added.
    """
    fsh_path = FSH_DIR / fsh_file
    if not fsh_path.exists():
        print(f"  WARNING: FSH file not found: {fsh_file}")
        return 0

    text = fsh_path.read_text()
    lines = text.split("\n")
    new_lines: list[str] = []
    bindings_added = 0

    for i, line in enumerate(lines):
        new_lines.append(line)

        # Match field definition lines like: * CD_FIELD 0..1 code "Label"
        m = re.match(r"^(\* )(\S+)\s+(\d+\.\.\S+)\s+code\s+", line)
        if not m:
            continue

        field_name = m.group(2)
        if field_name not in mappings:
            continue

        vs_name = mappings[field_name]
        binding_line = f"* {field_name} from {vs_name} (required)"

        # Skip if binding already exists on the next line
        if i + 1 < len(lines) and lines[i + 1].strip().startswith(f"* {field_name} from "):
            continue

        new_lines.append(binding_line)
        bindings_added += 1

    if bindings_added > 0:
        fsh_path.write_text("\n".join(new_lines))

    return bindings_added


def main() -> None:
    vs_ids = load_valueset_ids()
    print(f"Loaded {len(vs_ids)} ValueSet IDs")

    # Build configs from explicit sheet map
    all_configs: list[tuple[str, str, int, int, int, int, str]] = []
    for excel_file, sheet_name, fsh_file, layout in SHEET_MAP:
        if layout == "ortho":
            all_configs.append(build_orthopride_config(excel_file, sheet_name, fsh_file))
        elif layout == "angio":
            all_configs.append(build_angio_config(excel_file, sheet_name, fsh_file))
        else:
            all_configs.append(build_pacemaker_config(excel_file, sheet_name, fsh_file))

    total_bindings = 0
    total_misses = 0

    for config in all_configs:
        fsh_file = config[6]
        fsh_path = FSH_DIR / fsh_file
        if not fsh_path.exists():
            print(f"\nSKIP: {fsh_file} (not found)")
            continue

        mappings = extract_mappings(config, vs_ids)

        # Count code fields without mappings
        text = fsh_path.read_text()
        code_fields = re.findall(r"^\* (\S+)\s+\d+\.\.\S+\s+code\s+", text, re.MULTILINE)
        unmapped = [f for f in code_fields if f not in mappings]

        added = add_bindings_to_fsh(fsh_file, mappings)
        total_bindings += added
        total_misses += len(unmapped)

        status = "OK" if not unmapped else f"{len(unmapped)} unmapped"
        print(f"\n{fsh_file}: +{added} bindings ({status})")
        if unmapped:
            for f in unmapped:
                print(f"  UNMAPPED: {f}")

    print(f"\n{'='*60}")
    print(f"Total bindings added: {total_bindings}")
    print(f"Total unmapped code fields: {total_misses}")


if __name__ == "__main__":
    main()
